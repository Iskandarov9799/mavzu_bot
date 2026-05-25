import logging, io
from aiogram import Router, F, Bot
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from aiogram.fsm.context import FSMContext

from database.db import (
    get_all_users, get_user, get_full_stats,
    ban_user, unban_user, add_question,
    get_questions_page, search_questions, count_questions,
    delete_question, delete_bolim_questions,
    get_question_by_id, update_question,
    reset_all_subscriptions, import_questions_from_excel
)
from keyboards.keyboards import (
    admin_keyboard, main_menu_keyboard,
    subject_select_keyboard, bolim_select_keyboard,
    correct_answer_keyboard, member_action_keyboard,
    cancel_keyboard,
    qedit_page_keyboard, qedit_action_keyboard, qedit_fields_keyboard
)
from states import AdminStates, EditQuestionStates
from config import config

router  = Router()
logger  = logging.getLogger(__name__)
PAGE_SZ = 10

def is_admin(uid: int) -> bool:
    return uid in config.ADMIN_IDS

# ══════════════════════════════════════════════
# STATISTIKA
# ══════════════════════════════════════════════

@router.message(F.text == "📊 Statistika")
async def admin_stats(message: Message):
    if not is_admin(message.from_user.id): return
    s = await get_full_stats()
    await message.answer(
        f"📊 <b>Statistika</b>\n\n"
        f"👥 Foydalanuvchilar: <b>{s['users']}</b>\n"
        f"❓ Savollar: <b>{s['questions']}</b>\n"
        f"📝 Testlar: <b>{s['results']}</b>\n"
        f"💰 Kutayotgan: <b>{s['pending']}</b>",
        parse_mode="HTML"
    )

# ══════════════════════════════════════════════
# FOYDALANUVCHILAR
# ══════════════════════════════════════════════

@router.message(F.text == "👥 Foydalanuvchilar")
async def admin_users(message: Message):
    if not is_admin(message.from_user.id): return
    users = await get_all_users(50)
    text  = f"👥 <b>Foydalanuvchilar:</b>\n\n"
    for i, u in enumerate(users, 1):
        ban = "🚫 " if u.is_banned else ""
        text += f"{i}. {ban}<b>{u.full_name or '?'}</b> | {u.phone_number or '—'}\n"
    await message.answer(text, parse_mode="HTML")

@router.message(F.text == "👥 A'zolar")
async def members_list(message: Message):
    if not is_admin(message.from_user.id): return
    users = await get_all_users()
    total = len(users)
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    def mk_kb(page):
        buttons = []
        for u in users[page*PAGE_SZ:(page+1)*PAGE_SZ]:
            ban  = "🚫" if u.is_banned else "👤"
            name = u.full_name or "?"
            buttons.append([InlineKeyboardButton(
                text=f"{ban} {name} {u.phone_number or ''}",
                callback_data=f"member:{u.telegram_id}"
            )])
        nav = []
        if page > 0: nav.append(InlineKeyboardButton(text="◀️", callback_data=f"members:page:{page-1}"))
        nav.append(InlineKeyboardButton(text=f"{page+1}", callback_data="noop"))
        if (page+1)*PAGE_SZ < total: nav.append(InlineKeyboardButton(text="▶️", callback_data=f"members:page:{page+1}"))
        if nav: buttons.append(nav)
        buttons.append([InlineKeyboardButton(text="❌ Yopish", callback_data="members:close")])
        return InlineKeyboardMarkup(inline_keyboard=buttons)
    await message.answer(f"👥 <b>A'zolar</b> — jami {total} ta", reply_markup=mk_kb(0), parse_mode="HTML")

@router.callback_query(F.data.startswith("members:page:"))
async def members_page(callback: CallbackQuery):
    page  = int(callback.data.split(":")[2])
    users = await get_all_users()
    total = len(users)
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    buttons = []
    for u in users[page*PAGE_SZ:(page+1)*PAGE_SZ]:
        ban  = "🚫" if u.is_banned else "👤"
        buttons.append([InlineKeyboardButton(
            text=f"{ban} {u.full_name or '?'} {u.phone_number or ''}",
            callback_data=f"member:{u.telegram_id}"
        )])
    nav = []
    if page > 0: nav.append(InlineKeyboardButton(text="◀️", callback_data=f"members:page:{page-1}"))
    nav.append(InlineKeyboardButton(text=f"{page+1}", callback_data="noop"))
    if (page+1)*PAGE_SZ < total: nav.append(InlineKeyboardButton(text="▶️", callback_data=f"members:page:{page+1}"))
    if nav: buttons.append(nav)
    buttons.append([InlineKeyboardButton(text="❌ Yopish", callback_data="members:close")])
    try:
        await callback.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    except Exception: pass
    await callback.answer()

@router.callback_query(F.data.startswith("member:ban:"))
async def toggle_ban(callback: CallbackQuery, bot: Bot):
    tid  = int(callback.data.split(":")[2])
    user = await get_user(tid)
    if not user: await callback.answer("Topilmadi!"); return
    if user.is_banned:
        await unban_user(tid)
        await callback.answer("✅ Blok ochildi!")
        try: await bot.send_message(tid, "✅ Botdan blokinggiz ochildi.")
        except Exception: pass
    else:
        await ban_user(tid)
        await callback.answer("🚫 Bloklandi!")
    user = await get_user(tid)
    try:
        await callback.message.edit_reply_markup(reply_markup=member_action_keyboard(tid, user.is_banned))
    except Exception: pass

@router.callback_query(F.data.startswith("member:"))
async def member_detail(callback: CallbackQuery):
    if "ban" in callback.data: return
    tid  = int(callback.data.split(":")[1])
    user = await get_user(tid)
    if not user: await callback.answer("Topilmadi!"); return
    from database.db import get_user_results, get_subscription
    results = await get_user_results(tid)
    sub     = await get_subscription(tid)
    sub_txt = f"✅ {str(sub.expires_at)[:16]}" if sub else "❌ Yo'q"
    await callback.message.answer(
        f"👤 <b>{user.full_name or '?'}</b>\n"
        f"📱 {user.phone_number or '—'}\n"
        f"🆔 {user.telegram_id}\n"
        f"💳 Obuna: {sub_txt}\n"
        f"📝 Testlar: {len(results)}\n"
        f"{'🚫 Bloklangan' if user.is_banned else '✅ Faol'}",
        reply_markup=member_action_keyboard(tid, user.is_banned),
        parse_mode="HTML"
    )
    await callback.answer()

@router.callback_query(F.data == "members:close")
async def members_close(callback: CallbackQuery):
    try: await callback.message.delete()
    except Exception: pass
    await callback.answer()

@router.callback_query(F.data == "members:back")
async def members_back(callback: CallbackQuery):
    try: await callback.message.delete()
    except Exception: pass
    await callback.answer()

# ══════════════════════════════════════════════
# BROADCAST
# ══════════════════════════════════════════════

@router.message(F.text == "📢 Broadcast")
async def broadcast_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    await message.answer("📢 Xabarni yozing (matn/rasm/video):\n\n/cancel — bekor qilish")
    await state.set_state(AdminStates.broadcast)

@router.message(AdminStates.broadcast, F.text == "/cancel")
async def broadcast_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Bekor qilindi.", reply_markup=admin_keyboard())

@router.message(AdminStates.broadcast)
async def broadcast_send(message: Message, state: FSMContext, bot: Bot):
    await state.clear()
    users = await get_all_users()
    sent  = 0
    for u in users:
        if u.is_banned: continue
        try: await message.copy_to(u.telegram_id); sent += 1
        except Exception: pass
    await message.answer(f"✅ Yuborildi: <b>{sent}</b> ta", reply_markup=admin_keyboard(), parse_mode="HTML")

# ══════════════════════════════════════════════
# SAVOL QO'SHISH
# ══════════════════════════════════════════════

@router.message(F.text == "➕ Savol qo'shish")
async def addq_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    await state.clear()
    await message.answer("📚 Fanni tanlang:", reply_markup=subject_select_keyboard())

@router.callback_query(F.data.startswith("addq:subject:"))
async def addq_subject(callback: CallbackQuery, state: FSMContext):
    subject = callback.data.split(":")[2]
    await state.update_data(subject=subject)
    await callback.message.edit_text(
        f"📌 Bo'limni tanlang (0=Aralash, 1-40=Bo'limlar):",
        reply_markup=bolim_select_keyboard(subject)
    )
    await state.set_state(AdminStates.add_q_bolim)
    await callback.answer()

@router.callback_query(F.data.startswith("addq:bolim:"))
async def addq_bolim(callback: CallbackQuery, state: FSMContext):
    parts   = callback.data.split(":")
    subject = parts[2]
    bolim   = int(parts[3])
    await state.update_data(subject=subject, bolim=bolim)
    SUBJ = {"onatili": "📚 Ona tili", "adabiyot": "📖 Adabiyot"}
    bolim_txt = f"{bolim}-bo'lim" if bolim > 0 else "Aralash"
    await callback.message.edit_text(
        f"✏️ <b>{SUBJ[subject]} — {bolim_txt}</b>\n\nSavol matnini yozing:",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_q_text)
    await callback.answer()

@router.callback_query(F.data == "addq:cancel")
async def addq_cancel(message_or_cb, state: FSMContext):
    await state.clear()
    if isinstance(message_or_cb, CallbackQuery):
        try: await message_or_cb.message.delete()
        except Exception: pass
        await message_or_cb.bot.send_message(
            message_or_cb.from_user.id, "❌ Bekor qilindi.", reply_markup=admin_keyboard()
        )
        await message_or_cb.answer()
    else:
        await message_or_cb.answer("❌ Bekor qilindi.", reply_markup=admin_keyboard())

@router.message(AdminStates.add_q_text)
async def addq_text(message: Message, state: FSMContext):
    if message.text and message.text in ("❌ Bekor qilish",):
        await state.clear()
        await message.answer("❌ Bekor qilindi.", reply_markup=admin_keyboard()); return
    await state.update_data(text=message.text or "")
    await message.answer("🅰 A variantini yozing:")
    await state.set_state(AdminStates.add_q_a)

@router.message(AdminStates.add_q_a)
async def addq_a(message: Message, state: FSMContext):
    await state.update_data(a=message.text or "")
    await message.answer("🅱 B variantini yozing:")
    await state.set_state(AdminStates.add_q_b)

@router.message(AdminStates.add_q_b)
async def addq_b(message: Message, state: FSMContext):
    await state.update_data(b=message.text or "")
    await message.answer("🅲 C variantini yozing:")
    await state.set_state(AdminStates.add_q_c)

@router.message(AdminStates.add_q_c)
async def addq_c(message: Message, state: FSMContext):
    await state.update_data(c=message.text or "")
    await message.answer("🅳 D variantini yozing:")
    await state.set_state(AdminStates.add_q_d)

@router.message(AdminStates.add_q_d)
async def addq_d(message: Message, state: FSMContext):
    await state.update_data(d=message.text or "")
    await message.answer("✅ To'g'ri javobni tanlang:", reply_markup=correct_answer_keyboard())
    await state.set_state(AdminStates.add_question)

@router.callback_query(F.data.startswith("addq:correct:"))
async def addq_correct(callback: CallbackQuery, state: FSMContext):
    correct = callback.data.split(":")[2]
    await state.update_data(correct=correct)
    await callback.message.edit_text(
        "🖼 Rasm yuborish (ixtiyoriy):\n\n"
        "Rasm yuklamasa ⏭ bosing."
    )
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    await callback.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⏭ Rasmisiz saqlash", callback_data="addq:save")]
    ]))
    await state.set_state(AdminStates.add_q_image)
    await callback.answer()

@router.callback_query(F.data == "addq:save")
async def addq_save_no_image(callback: CallbackQuery, state: FSMContext):
    await _save_question(callback.message, state, None)
    await callback.answer()

@router.message(AdminStates.add_q_image, F.photo)
async def addq_image(message: Message, state: FSMContext):
    image_id = message.photo[-1].file_id
    await _save_question(message, state, image_id)

@router.message(AdminStates.add_q_image)
async def addq_no_image(message: Message, state: FSMContext):
    await _save_question(message, state, None)

async def _save_question(message, state: FSMContext, image_id: str = None):
    data = await state.get_data()
    await state.clear()
    qid = await add_question(
        subject=data['subject'], bolim=data['bolim'],
        text=data['text'],
        a=data.get('a',''), b=data.get('b',''),
        c=data.get('c',''), d=data.get('d',''),
        correct=data['correct'], image_id=image_id
    )
    SUBJ = {"onatili": "📚 Ona tili", "adabiyot": "📖 Adabiyot"}
    bolim_txt = f"{data['bolim']}-bo'lim" if data['bolim'] > 0 else "Aralash"
    await message.answer(
        f"✅ <b>Savol #{qid} saqlandi!</b>\n"
        f"{SUBJ[data['subject']]} | {bolim_txt}",
        reply_markup=admin_keyboard(), parse_mode="HTML"
    )

# ══════════════════════════════════════════════
# SAVOLLAR RO'YXATI
# ══════════════════════════════════════════════

# ══════════════════════════════════════════════
# SAVOLLAR MUHARRIRI
# ══════════════════════════════════════════════

PAGE_SZ_Q = 5

def _qfull(q) -> str:
    SUBJ = {"onatili": "📚 Ona tili", "adabiyot": "📖 Adabiyot"}
    bolim_txt = f"{q.bolim}-bo'lim" if q.bolim > 0 else "Aralash"
    return (
        f"🆔 ID: <b>{q.id}</b>\n"
        f"📚 Fan: <b>{SUBJ.get(q.subject, q.subject)}</b>\n"
        f"📌 Bo'lim: <b>{bolim_txt}</b>\n\n"
        f"❓ <b>Savol:</b>\n{q.question_text}\n\n"
        f"🅰 {q.option_a or '—'}\n"
        f"🅱 {q.option_b or '—'}\n"
        f"🅲 {q.option_c or '—'}\n"
        f"🅳 {q.option_d or '—'}\n\n"
        f"✅ To'g'ri: <b>{q.correct_answer or '—'}</b>"
        + (f"\n🖼 Rasm: bor" if q.image_file_id else "")
    )

@router.message(F.text == "📋 Savollar")
async def questions_list(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    await state.clear()
    prefix = "all"
    total  = await count_questions()
    qs     = await get_questions_page(offset=0, limit=PAGE_SZ_Q)
    if not qs:
        await message.answer("❌ Bazada savollar yo'q!"); return
    await message.answer(
        f"📋 <b>Savollar bazasi</b> — jami <b>{total}</b> ta\n\nSavolni bosib ko'ring:",
        reply_markup = qedit_page_keyboard(qs, 0, total, prefix),
        parse_mode   = "HTML"
    )
    await state.set_state(EditQuestionStates.browsing)

@router.callback_query(F.data.startswith("qedit:page:"))
async def qedit_turn_page(callback: CallbackQuery):
    parts  = callback.data.split(":")
    page   = int(parts[2])
    prefix = parts[3]

    if prefix.startswith("srch|"):
        keyword = prefix[5:]
        qs      = await search_questions(keyword)
        total   = len(qs)
        page_qs = qs[page*PAGE_SZ_Q:(page+1)*PAGE_SZ_Q]
    else:
        page_qs = await get_questions_page(offset=page*PAGE_SZ_Q, limit=PAGE_SZ_Q)
        total   = await count_questions()

    text = f"📋 <b>Savollar</b> — jami <b>{total}</b> ta\n\nSavolni bosing:"
    kb   = qedit_page_keyboard(page_qs, page, total, prefix)
    try:
        await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML")
    except Exception:
        try: await callback.message.delete()
        except Exception: pass
        await callback.bot.send_message(callback.from_user.id, text, reply_markup=kb, parse_mode="HTML")
    await callback.answer()

@router.callback_query(F.data.startswith("qedit:view:"))
async def qedit_view(callback: CallbackQuery):
    parts  = callback.data.split(":")
    qid    = int(parts[2])
    page   = int(parts[3])
    prefix = parts[4]
    q = await get_question_by_id(qid)
    if not q:
        await callback.answer("❌ Savol topilmadi!", show_alert=True); return
    text = _qfull(q)
    kb   = qedit_action_keyboard(qid, page, prefix)
    if q.image_file_id:
        try:
            await callback.message.delete()
            await callback.bot.send_photo(
                chat_id=callback.from_user.id, photo=q.image_file_id,
                caption=text, reply_markup=kb, parse_mode="HTML"
            )
        except Exception:
            try: await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML")
            except Exception: pass
    else:
        try: await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML")
        except Exception: pass
    await callback.answer()

@router.callback_query(F.data.startswith("qedit:edit:"))
async def qedit_edit(callback: CallbackQuery, state: FSMContext):
    parts  = callback.data.split(":")
    qid    = int(parts[2])
    page   = int(parts[3])
    prefix = parts[4]
    await state.update_data(edit_qid=qid, edit_page=page, edit_prefix=prefix)
    try:
        await callback.message.edit_text(
            f"✏️ <b>Savol #{qid}</b> — qaysi maydonni o'zgartirasiz?",
            reply_markup=qedit_fields_keyboard(qid, page, prefix), parse_mode="HTML"
        )
    except Exception:
        try: await callback.message.delete()
        except Exception: pass
        await callback.bot.send_message(
            callback.from_user.id,
            f"✏️ <b>Savol #{qid}</b> — qaysi maydonni o'zgartirasiz?",
            reply_markup=qedit_fields_keyboard(qid, page, prefix), parse_mode="HTML"
        )
    await state.set_state(EditQuestionStates.edit_field)
    await callback.answer()

@router.callback_query(F.data.startswith("qedit:field:"))
async def qedit_field_chosen(callback: CallbackQuery, state: FSMContext):
    parts  = callback.data.split(":")
    field  = parts[2]
    qid    = int(parts[3])
    page   = int(parts[4])
    prefix = parts[5]
    FIELDS = {
        "question_text":  "savol matni",
        "option_a":       "A varianti",
        "option_b":       "B varianti",
        "option_c":       "C varianti",
        "option_d":       "D varianti",
        "correct_answer": "to'g'ri javob (A/B/C/D)",
        "bolim":          "bo'lim raqami (0-40)",
        "subject":        "fan (onatili/adabiyot)",
        "image_file_id":  "rasm file_id",
    }
    await state.update_data(edit_qid=qid, edit_field=field, edit_page=page, edit_prefix=prefix)
    await callback.message.edit_text(
        f"✏️ <b>#{qid} — {FIELDS.get(field, field)}</b>\n\nYangi qiymatni yozing:",
        parse_mode="HTML"
    )
    await state.set_state(EditQuestionStates.edit_value)
    await callback.answer()

@router.message(EditQuestionStates.edit_value)
async def qedit_value_received(message: Message, state: FSMContext):
    data   = await state.get_data()
    qid    = data["edit_qid"]
    field  = data["edit_field"]
    page   = data.get("edit_page", 0)
    prefix = data.get("edit_prefix", "all")
    value  = message.text.strip() if message.text else ""
    if field == "bolim":
        try: value = int(value)
        except ValueError:
            await message.answer("❌ Raqam kiriting (0-40)!"); return
    await update_question(qid, **{field: value})
    await state.clear()
    q = await get_question_by_id(qid)
    await message.answer(
        f"✅ <b>Savol #{qid} yangilandi!</b>\n\n{_qfull(q)}",
        reply_markup=qedit_action_keyboard(qid, page, prefix),
        parse_mode="HTML"
    )

@router.callback_query(F.data.startswith("qedit:del:"))
async def qedit_delete_confirm(callback: CallbackQuery, state: FSMContext):
    parts  = callback.data.split(":")
    qid    = int(parts[2])
    page   = int(parts[3])
    prefix = parts[4]
    await state.update_data(del_qid=qid, del_page=page, del_prefix=prefix)
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Ha, o'chir", callback_data=f"qedit:delok:{qid}:{page}:{prefix}"),
        InlineKeyboardButton(text="❌ Yo'q",       callback_data=f"qedit:view:{qid}:{page}:{prefix}"),
    ]])
    try:
        await callback.message.edit_text(
            f"🗑 <b>Savol #{qid} ni o'chirasizmi?</b>\n⚠️ Qaytarib bo'lmaydi!",
            reply_markup=kb, parse_mode="HTML"
        )
    except Exception:
        try: await callback.message.delete()
        except Exception: pass
        await callback.bot.send_message(
            callback.from_user.id,
            f"🗑 <b>Savol #{qid} ni o'chirasizmi?</b>",
            reply_markup=kb, parse_mode="HTML"
        )
    await state.set_state(EditQuestionStates.confirm_delete)
    await callback.answer()

@router.callback_query(F.data.startswith("qedit:delok:"))
async def qedit_delete_confirmed(callback: CallbackQuery, state: FSMContext):
    parts  = callback.data.split(":")
    qid    = int(parts[2])
    page   = int(parts[3])
    prefix = parts[4]
    await delete_question(qid)
    await state.clear()
    total   = await count_questions()
    safe_pg = min(page, max(0, (total - 1) // PAGE_SZ_Q))
    qs      = await get_questions_page(offset=safe_pg*PAGE_SZ_Q, limit=PAGE_SZ_Q)
    await callback.message.edit_text(
        f"✅ <b>Savol #{qid} o'chirildi!</b>\n\n📋 Savollar — jami <b>{total}</b> ta:",
        reply_markup=qedit_page_keyboard(qs, safe_pg, total, prefix),
        parse_mode="HTML"
    )
    await callback.answer("✅ O'chirildi!")

@router.callback_query(F.data == "qedit:search")
async def qedit_search_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text(
        "🔍 <b>Qidirish</b>\n\nSavol matnidan kalit so'z yozing:",
        parse_mode="HTML"
    )
    await state.set_state(EditQuestionStates.searching)
    await callback.answer()

@router.message(EditQuestionStates.searching)
async def qedit_search_handler(message: Message, state: FSMContext):
    keyword = (message.text or "").strip()[:30]
    qs      = await search_questions(keyword)
    if not qs:
        await message.answer(f"🔍 '{keyword}' bo'yicha hech narsa topilmadi."); return
    await state.clear()
    prefix = f"srch|{keyword}"
    total  = len(qs)
    await message.answer(
        f"🔍 <b>'{keyword}'</b> — <b>{total}</b> ta natija:",
        reply_markup=qedit_page_keyboard(qs[:PAGE_SZ_Q], 0, total, prefix),
        parse_mode="HTML"
    )
    await state.set_state(EditQuestionStates.browsing)

@router.callback_query(F.data == "qedit:close")
async def qedit_close(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    try: await callback.message.delete()
    except Exception: pass
    await callback.bot.send_message(
        callback.from_user.id, "📋 Savollar muharriri yopildi.", reply_markup=admin_keyboard()
    )
    await callback.answer()

# ══════════════════════════════════════════════
# BO'LIM O'CHIRISH
# ══════════════════════════════════════════════

@router.message(F.text == "🗑 Bo'lim o'chirish")
async def delete_bolim_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    await message.answer(
        "📚 Qaysi fanning bo'limini o'chirish?",
        reply_markup=subject_select_keyboard("delbq:subject")
    )

@router.callback_query(F.data.startswith("delbq:subject:"))
async def delete_bolim_subject(callback: CallbackQuery, state: FSMContext):
    subject = callback.data.split(":")[2]
    await state.update_data(del_subject=subject)
    await callback.message.edit_text(
        "Bo'lim raqamini tanlang:",
        reply_markup=bolim_select_keyboard(subject, "delbq:bolim")
    )
    await state.set_state(AdminStates.delete_bolim)
    await callback.answer()

@router.callback_query(F.data.startswith("delbq:bolim:"))
async def delete_bolim_confirm(callback: CallbackQuery, state: FSMContext):
    parts   = callback.data.split(":")
    subject = parts[2]
    bolim   = int(parts[3])
    await state.clear()
    cnt = await count_questions(subject=subject, bolim=bolim)
    await delete_bolim_questions(subject=subject, bolim=bolim)
    SUBJ = {"onatili": "📚 Ona tili", "adabiyot": "📖 Adabiyot"}
    bolim_txt = f"{bolim}-bo'lim" if bolim > 0 else "Aralash"
    await callback.message.edit_text(
        f"✅ <b>{SUBJ[subject]} — {bolim_txt}</b>\n{cnt} ta savol o'chirildi.",
        parse_mode="HTML"
    )
    await callback.bot.send_message(callback.from_user.id, "Admin panel:", reply_markup=admin_keyboard())
    await callback.answer()

# ══════════════════════════════════════════════
# EXCEL IMPORT / EKSPORT
# ══════════════════════════════════════════════

def _make_shablon_bytes() -> bytes:
    """Shablon xlsx yaratish."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Savollar"

    HEADER_BG = "1e3a5f"; HEADER_FG = "ffffff"
    ROW1_BG   = "f0f4fa"; ROW2_BG   = "ffffff"
    HINT_BG   = "fff3cd"; HINT_FG   = "856404"
    thin   = Side(style="thin", color="c0c0c0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLUMNS = [
        ("subject", 18, "onatili yoki adabiyot"),
        ("bolim",   10, "0=Aralash, 1-40"),
        ("text",    60, "Savol matni"),
        ("a",       28, "A varianti"),
        ("b",       28, "B varianti"),
        ("c",       28, "C varianti"),
        ("d",       28, "D varianti"),
        ("correct", 12, "A, B, C yoki D"),
        ("image",   20, "Rasm file_id (ixtiyoriy)"),
    ]

    # 1-qator sarlavhalar
    for ci, (name, width, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, ci, name)
        cell.font      = Font(bold=True, color=HEADER_FG, name="Arial", size=11)
        cell.fill      = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 25

    # 2-qator izohlar
    for ci, (_, _, hint) in enumerate(COLUMNS, 1):
        cell = ws.cell(2, ci, hint)
        cell.font      = Font(italic=True, color=HINT_FG, name="Arial", size=9)
        cell.fill      = PatternFill("solid", start_color=HINT_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[2].height = 18

    # Namuna qatorlar
    SAMPLES = [
        ("onatili",  1, "O'zbek tilida nechta unli tovush bor?", "5 ta", "6 ta", "7 ta", "8 ta", "B", ""),
        ("onatili",  1, "Qaysi so'z ot turkumiga kiradi?", "yugurmoq", "baland", "maktab", "tez", "C", ""),
        ("onatili",  0, "Qaysi qatorda imlo xatosi yo'q?", "kitobxona", "kutubhona", "maktap", "daraxt'", "A", ""),
        ("adabiyot", 1, "Alisher Navoiy qaysi asrda yashagan?", "XIV asr", "XV asr", "XVI asr", "XVII asr", "B", ""),
        ("adabiyot", 1, "'Xamsa' asarining muallifi kim?", "Firdavsiy", "Sa'diy", "Navoiy", "Jomiy", "C", ""),
        ("adabiyot", 2, "'O'tkan kunlar' romani muallifi kim?", "Cho'lpon", "Fitrat", "Abdulla Qodiriy", "G'afur G'ulom", "C", ""),
    ]
    for ri, sample in enumerate(SAMPLES, 3):
        bg = ROW1_BG if ri % 2 == 1 else ROW2_BG
        for ci, val in enumerate(sample, 1):
            cell = ws.cell(ri, ci, val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A3"

    # Yo'riqnoma sheet
    ws2 = wb.create_sheet("Yo'riqnoma")
    INFO = [
        ("MAYDON", "QIYMAT", "IZOH"),
        ("subject", "onatili", "Ona tili fani"),
        ("subject", "adabiyot", "Adabiyot fani"),
        ("bolim", "0", "Barcha bo'limlardan aralash"),
        ("bolim", "1 dan 40 gacha", "Aniq bo'lim raqami"),
        ("correct", "A, B, C yoki D", "KATTA harf bilan yozing"),
        ("image", "(bo'sh)", "Rasm yo'q bo'lsa bo'sh qoldiring"),
    ]
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 40
    for ri, row in enumerate(INFO, 1):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(ri, ci, val)
            cell.border = border
            is_hdr = ri == 1
            cell.font = Font(bold=is_hdr, color=HEADER_FG if is_hdr else "000000", name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=HEADER_BG if is_hdr else ("f0f4fa" if ri%2 else "ffffff"))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws2.row_dimensions[ri].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.message(F.text == "📤 Excel import")
async def excel_import_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    shablon = _make_shablon_bytes()
    await message.answer_document(
        document = BufferedInputFile(shablon, filename="savollar_shablon.xlsx"),
        caption  = (
            "📤 <b>Excel import</b>\n\n"
            "Shu shablonni yuklab oling, to'ldiring va qayta yuboring.\n\n"
            "<b>Ustunlar:</b>\n"
            "• <code>subject</code> — onatili yoki adabiyot\n"
            "• <code>bolim</code> — 0 (aralash) yoki 1-40\n"
            "• <code>text</code> — savol matni\n"
            "• <code>a, b, c, d</code> — variantlar\n"
            "• <code>correct</code> — A, B, C yoki D\n"
            "• <code>image</code> — bo'sh qoldirilsa ham bo'ladi\n\n"
            "To'ldirilgan faylni yuboring:"
        ),
        parse_mode = "HTML"
    )
    await state.set_state(AdminStates.excel_import)

@router.message(AdminStates.excel_import, F.document)
async def excel_import_file(message: Message, state: FSMContext, bot: Bot):
    await state.clear()
    file    = await bot.get_file(message.document.file_id)
    data    = await bot.download_file(file.file_path)
    content = data.read()

    try:
        import openpyxl
        wb      = openpyxl.load_workbook(io.BytesIO(content))
        ws      = wb.active
        rows    = []
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
        # 2-qator izoh qatori bo'lishi mumkin — uni o'tkazib yuboramiz
        start_row = 3 if ws.max_row > 2 and not str(ws.cell(2, 1).value or "").strip().lower() in ("onatili","adabiyot") else 2
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if not any(row): continue
            r = {}
            for i, h in enumerate(headers):
                r[h] = row[i] if i < len(row) else None
            rows.append(r)
        added, errors = await import_questions_from_excel(rows)
        text = f"✅ <b>Import tugadi!</b>\n\n📥 Qo'shildi: <b>{added}</b> ta\n"
        if errors:
            err_txt = "\n".join(errors[:10])
            if len(errors) > 10:
                err_txt += f"\n... va yana {len(errors)-10} ta xato"
            text += f"⚠️ Xatolar ({len(errors)} ta):\n<code>{err_txt}</code>"
        await message.answer(text, reply_markup=admin_keyboard(), parse_mode="HTML")
    except Exception as e:
        await message.answer(f"❌ Import xato: {e}", reply_markup=admin_keyboard())

@router.message(AdminStates.excel_import)
async def excel_import_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Bekor qilindi.", reply_markup=admin_keyboard())

@router.message(F.text == "📥 Excel eksport")
async def excel_export(message: Message, bot: Bot):
    if not is_admin(message.from_user.id): return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from database.db import get_questions_page
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Savollar"
        headers = ["id", "subject", "bolim", "text", "a", "b", "c", "d", "correct", "image"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial")
            cell.fill = PatternFill("solid", start_color="1e3a5f")
            cell.font = Font(bold=True, color="ffffff", name="Arial")
        offset = 0
        while True:
            qs = await get_questions_page(offset=offset, limit=100)
            if not qs: break
            for q in qs:
                ws.append([
                    q.id, q.subject, q.bolim, q.question_text,
                    q.option_a, q.option_b, q.option_c, q.option_d,
                    q.correct_answer, q.image_file_id or ""
                ])
            offset += 100
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        await message.answer_document(
            document = BufferedInputFile(buf.read(), filename="questions_export.xlsx"),
            caption  = "📥 Savollar eksporti"
        )
    except Exception as e:
        await message.answer(f"❌ Eksport xato: {e}")

# ══════════════════════════════════════════════
# YECHIM LINKI
# ══════════════════════════════════════════════

@router.message(F.text == "🔗 Yechim linki")
async def solution_url_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    await message.answer(
        f"🔗 Hozirgi link: <code>{config.SOLUTION_URL or 'Yo\'q'}</code>\n\n"
        f"Yangi URL yozing:",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.set_solution_url)

@router.message(AdminStates.set_solution_url)
async def solution_url_set(message: Message, state: FSMContext):
    await state.clear()
    config.SOLUTION_URL = message.text.strip()
    await message.answer(
        f"✅ Yechim linki saqlandi:\n<code>{config.SOLUTION_URL}</code>",
        reply_markup=admin_keyboard(), parse_mode="HTML"
    )

# ══════════════════════════════════════════════
# TARIFLARNI NOLLASH
# ══════════════════════════════════════════════

@router.message(F.text == "♻️ Tariflarni nollash")
async def reset_tariffs(message: Message):
    if not is_admin(message.from_user.id): return
    await reset_all_subscriptions()
    await message.answer("✅ Barcha obunalar nollandi.", reply_markup=admin_keyboard())

# ══════════════════════════════════════════════
# NOOP
# ══════════════════════════════════════════════

@router.callback_query(F.data == "noop")
async def noop(callback: CallbackQuery):
    await callback.answer()